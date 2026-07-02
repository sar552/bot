"""Ma'lumotlarni Excel (.xlsx) fayliga eksport qilish (TZ §10).

Google Sheets'ga muqobil — admin bir tugma bosib, joriy holatni
Telegram chatida Excel fayl sifatida oladi. Tashqi sozlash kerak emas.
"""
from __future__ import annotations

import asyncio
import io

from openpyxl import Workbook
from openpyxl.styles import Font

from app.db.repositories import BonusRepo, CodeRepo, UserRepo, purchase_report
from app.db.session import session_factory

_USERS_HEADER = ["telegram_id", "username", "name", "phone", "registered_at"]
_CODES_HEADER = ["code", "status", "used_by_telegram_id", "name", "phone",
                 "username", "used_at", "note"]
_BONUSES_HEADER = ["bonus_code", "discount_percent", "status",
                   "assigned_to_telegram_id", "assigned_to_name",
                   "assigned_to_phone", "assigned_to_username",
                   "assigned_at", "used_at", "expires_at", "note"]
# Hisobot: bitta qatorda mijoz + ishlatgan kod + olgan bonus
_REPORT_HEADER = ["mijoz", "telefon", "username", "telegram_id",
                  "ishlatgan_kod", "kod_sanasi",
                  "bonus_kod", "bonus_status", "bonus_berilgan", "bonus_muddati"]


def _fmt(value) -> str | int:
    if value is None:
        return ""
    if hasattr(value, "value"):  # enum
        return value.value
    if hasattr(value, "isoformat"):  # datetime
        return value.isoformat(sep=" ", timespec="seconds")
    return value


def _write_sheet(ws, header: list[str], rows: list[list]) -> None:
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_fmt(v) for v in row])
    # Ustun kengligini taxminan moslash
    for i, _ in enumerate(header, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 20


async def build_export() -> bytes:
    """Postgres'dan joriy holatni o'qib, 3 varaqli .xlsx baytlarini qaytaradi."""
    async with session_factory() as session:
        users = await UserRepo(session).list_all(limit=100_000)
        codes = await CodeRepo(session).list_by_status(None, limit=100_000)
        bonuses = await BonusRepo(session).list_by_status(None, limit=100_000)
        report = await purchase_report(session, limit=100_000)

    def make() -> bytes:
        wb = Workbook()

        ws_users = wb.active
        ws_users.title = "Users"
        _write_sheet(ws_users, _USERS_HEADER, [
            [u.telegram_id, u.username, u.full_name, u.phone, u.registered_at]
            for u in users
        ])

        ws_codes = wb.create_sheet("Codes")
        _write_sheet(ws_codes, _CODES_HEADER, [
            [c.code, c.status, c.used_by_telegram_id, c.used_by_name,
             c.used_by_phone, c.used_by_username, c.used_at, c.note]
            for c in codes
        ])

        ws_bonuses = wb.create_sheet("Bonuses")
        _write_sheet(ws_bonuses, _BONUSES_HEADER, [
            [b.bonus_code, b.discount_percent, b.status, b.assigned_to_telegram_id,
             b.assigned_to_name, b.assigned_to_phone, b.assigned_to_username,
             b.assigned_at, b.used_at, b.expires_at, b.note]
            for b in bonuses
        ])

        # Hisobot: kim qaysi kodni ishlatib, qaysi bonusni olgani — bitta qatorda
        ws_report = wb.create_sheet("Hisobot", index=0)  # birinchi varaq
        _write_sheet(ws_report, _REPORT_HEADER, [list(r) for r in report])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # openpyxl bloklovchi — alohida thread'da bajaramiz
    return await asyncio.to_thread(make)
